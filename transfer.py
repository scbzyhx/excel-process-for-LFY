#
# -*- coding:utf-8 -*-
import logging
import os
from datetime import datetime
import json
import sys

from openpyxl import load_workbook

#COPY RIGHT
copy_right = "Copyright ownership belongs to YHX, shall not be reproduced , copied, or used in other ways without permission. \nOtherwise YHX will have the right to pursue legal responsibilities.\nMail to the Author: scbzyhx@gmail.com"

SHEET_3 = 'Master' #工作表格中取sheet3
INFO_SHEET = u'Sheet1' #信息表中取Sheet1
OUTPUT_SHEET = u'Master' #利用sheet3 中信息格式化输出
DIR = u"../"  #工作所在目录 
OUTPUT =  DIR + u"output/" #输出的路径
LOG = OUTPUT + u"log/" #日志路径
INFO_FILE = u"信息表.xlsx" #默认信息表

#这一部分是输出文件的名称
PRE_SEQ = "DS_SLCSZX_"+ datetime.now().strftime('%Y%m%d%H%M%S')
SEQ = 0
TAB = "\t"

#配置文件的名称
CONFIG_FILE = "config.txt"

#信息表中，相应信息所在的格子位置
"""
INFO TABLE
"""
NAMES = {
'HAWB' : 'A',
'CTRL' : 'B',
'SENDER_ID' : 'C',
'RECEIVER_ID' : 'D',
'MESSAGE_TYPE' : 'E',
'CONTROL_ID' : 'F',
'DATETIMESTAMP' : 'G',
'LANGCODE' : 'H',
'PARTNER_DELIVERY_DATE' : 'I',
'SHIPPED_DATE' : 'J',
'SHIPPED_TIME' : 'K',
'PICKUP_DATE':'L',
'PICKUP_TIME':'M',
'SHIPMENT_ID_OR_TRACKING_NUMBER' : 'N',
'PORT_OF_EXPORT' : 'O',
'PORT_OF_ENTRY' : 'P',
'NUM_OF_CARTONS' : 'Q'
}
ROW_LEN = [(chr(ord("A") + i)) for i in xrange(26)]
ROW_LEN.extend(['AA','AB','AC'])

#需要输出的格子的位置
CELL_PRINT1 = [(chr(ord("A") + i))+"9" for i in xrange(7)]
CELL_PRINT1.extend(["A1" for i in xrange(31-7)])
#print CELL_PRINT1
#CELL_PRINT2 = []
CELL_PRINT2 = [(chr(ord("A") + i))+"12" for i in xrange(26)]
CELL_PRINT2.extend(["AA12","AB12","AC12","AD12","AE12"])
#CELL_PRINT.extend()
CELL_PRINT = [CELL_PRINT1,CELL_PRINT2]
#print CELL_PRINT
OTHER_CELL_PRINT = [(chr(ord("A") + i)) for i in xrange(26)]
OTHER_CELL_PRINT.extend(["AA","AB","AC","AD","AE"])

#创建目录
if not os.path.isdir(OUTPUT):
    os.mkdir(OUTPUT)
if not os.path.isdir(LOG):
    os.mkdir(LOG)

#setup loggers
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

handler = logging.FileHandler(LOG+datetime.now().strftime('output_%H_%M_%d_%m_%Y.log'))
handler.setLevel(logging.INFO)

formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
handler.setFormatter(formatter)


logger.addHandler(handler)
#logger.info("hello world")

def pause():
    raw_input("Press Enter to stop")
    

class Empty_Sheet(Exception):
    def __init__(self,sheet_name = "Master"):
        self.name = sheet_name

class WorkBook(object):
    def __init__(self,file_name = None):
        super(WorkBook,self).__init__()
        self.wb = None
        self.sheet = None
        self.file = file_name
        
    def open(self,filename = None,password = None):
        if filename is None and self.file is None :
            raise Exception()
        if self.file is None:
            self.file = filename
        try:
            self.wb =  load_workbook(self.file,data_only=True)
        except IOError:
            logger.error("Failed to open: " +self.file)
            print "Failed to open: " +self.file
            pause()            
            sys.exit(-1)
        #self.sheet = self.wb.get_sheet_by_name(SHEET_3)

    def modify(self):
        if self.sheet is None:
            raise Empty_Sheet()

class InfoTable(WorkBook):
    def __init__(self,info_file = u"信息表.xlsx"):
        super(InfoTable,self).__init__(info_file)
        self.dicts = {}
        self.open()
        self.build_dict()
        
    def open(self,filename = None,password = None):
        super(InfoTable,self).open(filename,password)
        try:            
            self.sheet = self.wb.get_sheet_by_name(INFO_SHEET)
        except KeyError as k:
            print "There is No INFO_SHEET:" + INFO_SHEET + "  in file:" + self.file
            pause()
            sys.exit(-1)
        
    def build_dict(self):
        i = 2 ##start row
        while True:
            hawb = self.sheet[NAMES['HAWB']+str(i)]
            #print hawb.value
            if hawb.value == None:
                break
            hawb_value = hawb.value
            self.dicts.setdefault(hawb_value,{})
            for k,v in NAMES.items():
                self.dicts[hawb_value][k] = self.sheet[v+str(i)].value    
            i += 1
    def get(self,HAWB,COL):
        if COL is None or HAWB is None:
            return None
        #print HAWB
        #print COL
        #print self.dicts.keys()
        row = self.dicts.get(HAWB,{})
        #print row

        return row.get(COL,None)
            
    def show(self):
        for x in self.dicts.keys():
            print self.dicts[x]
        
class WorkTable(WorkBook):
    CELLS = {
        'HAWB_IN_WORK_TABLE' : 'L12',
        'CTRL':'A9',
        'SENDER_ID':'B9',
        'RECEIVER_ID':'C9',
        'MESSAGE_TYPE':'D9',
        'CONTROL_ID':'E9',
        'DATETIMESTAMP':'F9',
        'LANGCODE':'G9',

        'PARTNER_DELIVERY_DATE':'C12',
        'SHIPPED_DATE':'D12',
        'SHIPPED_TIME':'E12',
        'PICKUP_DATE':'F12',
        'PICKUP_TIME':'G12',
        'SHIPMENT_ID':'K12',
        'PORT_OF_EXPORT':'Q12',
        'PORT_OF_ENTRY':'R12',
        'NUM_OF_CARTONS':'AE12'
    }
    MAP_TO_NAME = {
        'HAWB_IN_WORK_TABLE' : 'HAWB',
        'CTRL':'CTRL',
        'SENDER_ID':'SENDER_ID',
        'RECEIVER_ID':'RECEIVER_ID',
        'MESSAGE_TYPE':'MESSAGE_TYPE',
        'CONTROL_ID':'CONTROL_ID',
        'DATETIMESTAMP':'DATETIMESTAMP',
        'LANGCODE':'LANGCODE',

        'PARTNER_DELIVERY_DATE':'PARTNER_DELIVERY_DATE',
        'SHIPPED_DATE':'SHIPPED_DATE',
        'SHIPPED_TIME':'SHIPPED_TIME',
        'PICKUP_DATE':'PICKUP_DATE',
        'PICKUP_TIME':'PICKUP_TIME',
        'SHIPMENT_ID':'SHIPMENT_ID_OR_TRACKING_NUMBER',
        'PORT_OF_EXPORT':'PORT_OF_EXPORT',
        'PORT_OF_ENTRY':'PORT_OF_ENTRY',
        'NUM_OF_CARTONS':'NUM_OF_CARTONS'
    }

    def __init__(self,info_table,file_name=None):
        super(WorkTable,self).__init__(file_name)
        self.info_table = info_table
        if file_name is not None:
            self.open()
    
    def open(self,filename = None,password = None):
        super(WorkTable,self).open(filename,password)
        
        self.modify_sheet = self.wb.get_sheet_by_name(SHEET_3)

            
    def get_hawb(self):
        return self.modify_sheet[WorkTable.CELLS['HAWB_IN_WORK_TABLE']].value
        
    def modify(self):
        hawb = self.get_hawb()
        
        for k,v in WorkTable.CELLS.items():
            #print k,"  is"
            #print s
            if k == "HAWB_IN_WORK_TABLE":
                continue
            self.modify_sheet[WorkTable.CELLS[k]] = self.info_table.get(hawb,WorkTable.MAP_TO_NAME[k])
            #print self.modify_sheet[WorkTable.CELLS[k]].value
        self.output_sheet = self.wb.get_sheet_by_name(OUTPUT_SHEET)
        #self.print_sheet()
    def print_sheet(self,fl):
        for row in CELL_PRINT:
            for i,cell in enumerate(row):
                cell_value = self.output_sheet[cell].value
                if cell_value == None:
                    cell_value = ""
                fl.write(str(cell_value))
                if i < len(row)-1:
                    #print "write tab"
                    fl.write(TAB)
            fl.write("\n")

        START = 15
        END = 1512
        #(chr(ord("A") + i))
        for st in xrange(START,END+1):
            for i in OTHER_CELL_PRINT:
                cell = i + str(st)#chr(ord("A") + i) + str(st)
                cell_value = self.output_sheet[cell].value
                #print cell_value
                if cell_value == None:
                    cell_value = ""
                
                
                fl.write(str(cell_value))
                if i !=  OTHER_CELL_PRINT[-1]:
                    fl.write(TAB)
            fl.write( "\n")
    def test_formula(self):
        self.sh = self.wb.get_sheet_by_name("Flatfile format")
        formula = self.sh["A3"]
        #print type(formula.internal_value)
        #print formula.value
        #print dir(formula)

        
        
            
                
        
    def __del__(self):
        pass
        #self.wb.save(filename="tmp.xlsx")
                
        
        

def test():
    infot = InfoTable(u"信息表.xlsx")
    #infot.open()
    #infot.
    b = WorkTable(infot,u"C5P21   856-原始的.xlsx")
    b.modify()

def get_file_name():
    global SEQ
    file_name = PRE_SEQ + ("%06d" % SEQ) + ".txt"
    SEQ += 1
    return file_name

def load_config_json():
    try:
        json_data = open(CONFIG_FILE).read()
    except Exception as e:
        print "No config file, use default setting, 目录为： ..  ,  信息表名称为：信息表.xlsx"
        return
    #print json_data
    data = json.loads(json_data)
    json.dumps(data)
    global DIR
    if data.has_key("DIR"):
        DIR = data["DIR"] + "/"
    global INFO_TABLE
    if data.has_key("INFO_FILE"):
        INFO_FILE = data["INFO_FILE"]
    global PRE_SEQ
    if data.has_key("PRE_SEQ"):
        PRE_SEQ = data["PRE_SEQ"]+ datetime.now().strftime('%Y%m%d%H%M%S')
    

def main():
    info_table = InfoTable(DIR+INFO_FILE)
    files = os.listdir(DIR)
    for fl in files:
        #print fl
        if os.path.isfile(DIR+fl) == False:
            continue
        filename = os.path.split(DIR+fl)
        #print filename[-1]
        if filename[-1].split('.')[-1] != "xlsx":
            #print filename
            continue
        #print filename[-1]
        try:
            wb = WorkTable(info_table,DIR+fl)
            #wb.test_formula()
            #return
            wb.modify()
            new_file_name = get_file_name()
            with open(OUTPUT + new_file_name,"w+")as f:
                #print "open success"
                wb.print_sheet(f)
                logger.info(filename[-1] + u" transfered to " + new_file_name)
                #print type(fl)
                print (filename[-1]+ u" is transfered")
                
            

        except Exception as e:
            
            logger.error(filename[-1]+" transfer failed! Exception is %s" % e)


if __name__ == "__main__":
    print copy_right
    print "\n"
    print "Start transfering...."
    load_config_json()
    main()
    print u"输出的文件在 output 文件夹中。\n"
    pause()
        
#x = InfoTable(u"信息表.xlsx")
#x.open()
#x.build_dict()
#x.show()
"""
b = WorkBook()
b.open(u"C5P21   856-原始的.xlsx")
print b.wb
print dir(b.sheet)
"""

#wb = load_workbook(filename='test.xlsx')

#sheet_ranges = wb.get_active_sheet()

#print(sheet_ranges['A1'].value)
